import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
import os
import uuid
from datetime import datetime
import openpyxl
from PIL import Image, ImageDraw, ImageFont
import re
from fpdf import FPDF

class UnexcaCertificateSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Certificados UNEXCA")
        self.root.geometry("1200x800")
        self.root.configure(bg='#ECEFF1')
        self.root.resizable(False, False)

        # Crear directorios necesarios
        self.crear_directorios()

        # Configuración de estilos
        self.configurar_estilos()

        # Inicializar base de datos
        self.inicializar_base_datos()

        # Crear interfaz principal
        self.crear_interfaz_principal()

    def crear_directorios(self):
        directorios = [
            "bases_datos", 
            "certificados", 
            "templates", 
            "fonts", 
            "logs",
            "exports"
        ]
        for directorio in directorios:
            os.makedirs(directorio, exist_ok=True)

    def configurar_estilos(self):
        self.colores = {
            'fondo_principal': '#ECEFF1',
            'texto_principal': '#2C3E50',
            'boton_primario': '#3498DB',
            'boton_secundario': '#2ECC71',
            'bordes': '#B0BEC5'
        }

        self.fuentes = {
            'titulo_grande': ('Helvetica', 24, 'bold'),
            'texto_normal': ('Arial', 12),
            'texto_pequeño': ('Arial', 10)
        }

        estilo = ttk.Style()
        estilo.theme_use('clam')
        
        estilo.configure('primary.TButton', 
                         background=self.colores['boton_primario'], 
                         foreground='white', 
                         font=self.fuentes['texto_normal'],
                         borderwidth=2,
                         relief="flat")
        
        estilo.configure('secondary.TButton', 
                         background=self.colores['boton_secundario'], 
                         foreground='white', 
                         font=self.fuentes['texto_normal'],
                         borderwidth=2,
                         relief="flat")
        
        estilo.configure('TFrame', background=self.colores['fondo_principal'])
        estilo.configure('TLabel', background=self.colores['fondo_principal'], foreground=self.colores['texto_principal'])
        estilo.configure('TEntry', bordercolor=self.colores['bordes'], relief="flat")

    def inicializar_base_datos(self):
        db_path = os.path.join('bases_datos', 'unexca_certificados.db')
        
        try:
            self.conn = sqlite3.connect(db_path)
            self.cursor = self.conn.cursor()

            self.cursor.executescript('''
                CREATE TABLE IF NOT EXISTS estudiantes (
                    id TEXT PRIMARY KEY,
                    nombre TEXT NOT NULL,
                    apellido TEXT NOT NULL,
                    cedula TEXT UNIQUE,
                    email TEXT,
                    fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS cursos (
                    id TEXT PRIMARY KEY,
                    nombre TEXT NOT NULL,
                    codigo TEXT UNIQUE,
                    area TEXT,
                    duracion TEXT,
                    descripcion TEXT,
                    instructor TEXT,
                    fecha_creacion DATETIME DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS certificados (
                    id TEXT PRIMARY KEY,
                    estudiante_id TEXT,
                    curso_id TEXT,
                    fecha_emision DATETIME,
                    archivo_certificado TEXT,
                    FOREIGN KEY(estudiante_id) REFERENCES estudiantes(id),
                    FOREIGN KEY(curso_id) REFERENCES cursos(id)
                );
            ''')
            
            self.conn.commit()
            
        except sqlite3.Error as e:
            messagebox.showerror("Error de Base de Datos", f"No se pudo inicializar la base de datos: {e}")
            raise

    def crear_interfaz_principal(self):
        frame_principal = ttk.Frame(self.root, style='TFrame')
        frame_principal.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        titulo = ttk.Label(
            frame_principal, 
            text="Sistema de Certificados UNEXCA",
            font=self.fuentes['titulo_grande'],
            style='TLabel'
        )
        titulo.pack(pady=20)

        modulos = [
            ("Gestión de Estudiantes", self.abrir_gestion_estudiantes),
            ("Gestión de Cursos", self.abrir_gestion_cursos),
            ("Generación de Certificados", self.abrir_generacion_certificados)
        ]

        for texto, comando in modulos:
            boton = ttk.Button(
                frame_principal, 
                text=texto, 
                command=comando,
                style='primary.TButton',
                width=40
            )
            boton.pack(pady=10)

    def abrir_gestion_estudiantes(self):
        ventana_estudiantes = tk.Toplevel(self.root)
        ventana_estudiantes.title("Gestión de Estudiantes")
        ventana_estudiantes.geometry("1000x600")
        ventana_estudiantes.configure(bg='#ECEFF1')

        # Frame para entrada de datos
        frame_entrada = ttk.Frame(ventana_estudiantes, style='TFrame')
        frame_entrada.pack(pady=10, padx=20, fill=tk.X)

        # Campos de entrada
        campos = [
            ("Nombre", "nombre"),
            ("Apellido", "apellido"),
            ("Cédula", "cedula"),
            ("Email", "email")
        ]

        entradas = {}
        for i, (label, campo) in enumerate(campos):
            ttk.Label(frame_entrada, text=label, style='TLabel').grid(row=i//2, column=(i%2)*2, padx=5, pady=5)
            entrada = ttk.Entry(frame_entrada, width=30, style='TEntry')
            entrada.grid(row=i//2, column=(i%2)*2+1, padx=5, pady=5)
            entradas[campo] = entrada

        # Botones de acción
        frame_botones = ttk.Frame(ventana_estudiantes, style='TFrame')
        frame_botones.pack(pady=10)

        botones = [
            ("Registrar", lambda: self.registrar_estudiante(entradas)),
            ("Cargar Estudiantes", lambda: self.cargar_estudiantes(tabla)),
            ("Importar Excel", self.importar_estudiantes_excel)
        ]

        for texto, comando in botones:
            ttk.Button(frame_botones, text=texto, command=comando, style='primary.TButton').pack(side=tk.LEFT, padx=5)

        # Tabla de estudiantes
        columns = ("ID", "Nombre", "Apellido", "Cédula", "Email")
        tabla = ttk.Treeview(ventana_estudiantes, columns=columns, show="headings")
        
        for col in columns:
            tabla.heading(col, text=col)
            tabla.column(col, width=150, anchor=tk.CENTER)
        
        tabla.pack(expand=True, fill=tk.BOTH, padx=20, pady=10)

        # Cargar estudiantes iniciales
        self.cargar_estudiantes(tabla)

    def registrar_estudiante(self, entradas):
        # Validaciones básicas
        datos = {}
        for campo, entrada in entradas.items():
            valor = entrada.get().strip()
            if not valor:
                messagebox.showerror("Error", f"El campo {campo} no puede estar vacío")
                return
            datos[campo] = valor

        # Validar email
        if not re.match(r"[^@]+@[^@]+\.[^@]+", datos['email']):
            messagebox.showerror("Error", "Email inválido")
            return

        # Generar ID único
        estudiante_id = str(uuid.uuid4())
        
        try:
            self.cursor.execute('''
                INSERT INTO estudiantes (id, nombre, apellido, cedula, email) 
                VALUES (?, ?, ?, ?, ?)
            ''', (estudiante_id, datos['nombre'], datos['apellido'], 
                  datos['cedula'], datos['email']))
            
            self.conn.commit()
            messagebox.showinfo("Éxito", "Estudiante registrado correctamente")
            
            # Limpiar entradas
            for entrada in entradas.values():
                entrada.delete(0, tk.END)
        
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "Ya existe un estudiante con esta cédula")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo registrar el estudiante: {e}")

    def cargar_estudiantes(self, tabla=None):
        try:
            self.cursor.execute("SELECT * FROM estudiantes")
            estudiantes = self.cursor.fetchall()

            if tabla:
                # Limpiar tabla
                for i in tabla.get_children():
                    tabla.delete(i)
                
                # Insertar estudiantes
                for estudiante in estudiantes:
                    tabla.insert("", "end", values=estudiante)
            
            return estudiantes
        
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar los estudiantes: {e}")

    def importar_estudiantes_excel(self):
        archivo = filedialog.askopenfilename(
            filetypes=[("Archivos Excel", "*.xlsx *.xls")]
        )
        
        if not archivo:
            return

        try:
            wb = openpyxl.load_workbook(archivo)
            hoja = wb.active

            for fila in hoja.iter_rows(min_row=2, values_only=True):
                if len(fila) >= 4:
                    nombre, apellido, cedula, email = fila[:4]
                    
                    estudiante_id = str(uuid.uuid4())
                    
                    try:
                        self.cursor.execute('''
                            INSERT INTO estudiantes (id, nombre, apellido, cedula, email) 
                            VALUES (?, ?, ?, ?, ?)
                        ''', (estudiante_id, nombre, apellido, cedula, email))
                    except sqlite3.IntegrityError:
                        # Omitir registros duplicados
                        continue

            self.conn.commit()
            messagebox.showinfo("Éxito", "Estudiantes importados correctamente")
        
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron importar los estudiantes: {e}")

    def abrir_gestion_cursos(self):
        ventana_cursos = tk.Toplevel(self.root)
        ventana_cursos.title("Gestión de Cursos")
        ventana_cursos.geometry("1000x600")
        ventana_cursos.configure(bg='#ECEFF1')

        # Frame para entrada de datos
        frame_entrada = ttk.Frame(ventana_cursos, style='TFrame')
        frame_entrada.pack(pady=10, padx=20, fill=tk.X)

        # Campos de entrada
        campos = [
            ("Nombre del Curso", "nombre"),
            ("Código del Curso", "codigo"),
            ("Área", "area"),
            ("Duración", "duracion"),
            ("Descripción", "descripcion"),
            ("Instructor", "instructor")
        ]

        entradas = {}
        for i, (label, campo) in enumerate(campos):
            ttk.Label(frame_entrada, text=label, style='TLabel').grid(row=i//2, column=(i%2)*2, padx=5, pady=5)
            entrada = ttk.Entry(frame_entrada, width=30, style='TEntry')
            entrada.grid(row=i//2, column=(i%2)*2+1, padx=5, pady=5)
            entradas[campo] = entrada

        # Botones de acción
        frame_botones = ttk.Frame(ventana_cursos, style='TFrame')
        frame_botones.pack(pady=10)

        botones = [
            ("Registrar Curso", lambda: self.registrar_curso(entradas)),
            ("Cargar Cursos", lambda: self.cargar_cursos(tabla))
        ]

        for texto, comando in botones:
            ttk.Button(frame_botones, text=texto, command=comando, style='primary.TButton').pack(side=tk.LEFT, padx=5)

        # Tabla de cursos
        columns = ("ID", "Nombre", "Código", "Área", "Duración", "Descripción", "Instructor")
        tabla = ttk.Treeview(ventana_cursos, columns=columns, show="headings")
        
        for col in columns:
            tabla.heading(col, text=col)
            tabla.column(col, width=150, anchor=tk.CENTER)
        
        tabla.pack(expand=True, fill=tk.BOTH, padx=20, pady=10)

        # Cargar cursos iniciales
        self.cargar_cursos(tabla)

    def registrar_curso(self, entradas):
        # Validaciones básicas
        datos = {}
        for campo, entrada in entradas.items():
            valor = entrada.get().strip()
            if not valor:
                messagebox.showerror("Error", f"El campo {campo} no puede estar vacío")
                return
            datos[campo] = valor

        # Generar ID único
        curso_id = str(uuid.uuid4())
        
        try:
            self.cursor.execute('''
                INSERT INTO cursos (id, nombre, codigo, area, duracion, descripcion, instructor) 
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (curso_id, datos['nombre'], datos['codigo'], datos['area'], 
                  datos['duracion'], datos['descripcion'], datos['instructor']))
            
            self.conn.commit()
            messagebox.showinfo("Éxito", "Curso registrado correctamente")
            
            # Limpiar entradas
            for entrada in entradas.values():
                entrada.delete(0, tk.END)
        
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "Ya existe un curso con este código")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo registrar el curso: {e}")

    def cargar_cursos(self, tabla=None):
        try:
            self.cursor.execute("SELECT * FROM cursos")
            cursos = self.cursor.fetchall()

            if tabla:
                # Limpiar tabla
                for i in tabla.get_children():
                    tabla.delete(i)
                
                # Insertar cursos
                for curso in cursos:
                    tabla.insert("", "end", values=curso)
            
            return cursos
        
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar los cursos: {e}")

    def abrir_generacion_certificados(self):
        ventana_certificados = tk.Toplevel(self.root)
        ventana_certificados.title("Generación de Certificados")
        ventana_certificados.geometry("600x400")
        ventana_certificados.configure(bg='#ECEFF1')

        # Frame para selección de curso
        frame_seleccion = ttk.Frame(ventana_certificados, style='TFrame')
        frame_seleccion.pack(pady=20, padx=20, fill=tk.X)

        ttk.Label(frame_seleccion, text="Seleccionar Curso:", style='TLabel').grid(row=0, column=0, padx=5, pady=5)
        cursos_combo = ttk.Combobox(frame_seleccion, state="readonly")
        cursos_combo.grid(row=0, column=1, padx=5, pady=5)

        # Cargar cursos en el combobox
        cursos = self.cargar_cursos()
        cursos_combo['values'] = [f"{curso[1]} ({curso[2]})" for curso in cursos]

        # Botón para generar certificados
        boton_generar = ttk.Button(
            ventana_certificados, 
            text="Generar Certificados", 
            command=lambda: self.generar_certificados(cursos_combo.get(), cursos),
            style='primary.TButton'
        )
        boton_generar.pack(pady=20)

    def generar_certificados(self, curso_seleccionado, cursos):
        if not curso_seleccionado:
            messagebox.showerror("Error", "Debe seleccionar un curso")
            return
        
        curso_id = None
        for curso in cursos:
            if f"{curso[1]} ({curso[2]})" == curso_seleccionado:
                curso_id = curso[0]
                break
        
        if not curso_id:
            messagebox.showerror("Error", "Curso no encontrado")
            return

        try:
            self.cursor.execute("SELECT * FROM estudiantes")
            estudiantes = self.cursor.fetchall()
            
            if not estudiantes:
                messagebox.showerror("Error", "No hay estudiantes registrados")
                return

            # Generar certificados en PDF
            for estudiante in estudiantes:
                self._generar_pdf(estudiante, curso_id)
            
            messagebox.showinfo("Éxito", "Certificados generados correctamente")
        
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron generar los certificados: {e}")

    def _generar_pdf(self, estudiante, curso_id):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        pdf.cell(200, 10, txt="CERTIFICADO DE PARTICIPACIÓN", ln=True, align='C')
        pdf.ln(10)
        
        pdf.cell(200, 10, txt=f"Certificamos que {estudiante[1]} {estudiante[2]}", ln=True, align='C')
        pdf.cell(200, 10, txt=f"con cédula {estudiante[3]}", ln=True, align='C')
        pdf.cell(200, 10, txt=f"ha completado satisfactoriamente el curso:", ln=True, align='C')
        pdf.ln(10)
        
        self.cursor.execute("SELECT nombre, codigo FROM cursos WHERE id = ?", (curso_id,))
        curso = self.cursor.fetchone()
        if curso:
            pdf.cell(200, 10, txt=f"{curso[0]} ({curso[1]})", ln=True, align='C')
        
        pdf.ln(20)
        pdf.cell(200, 10, txt=f"Fecha: {datetime.now().strftime('%Y-%m-%d')}", ln=True, align='C')

        # Guardar el PDF
        archivo_certificado = f"certificados/{estudiante[1]}_{estudiante[2]}_{curso[1]}.pdf"
        pdf.output(archivo_certificado)
        
        # Registrar el certificado en la base de datos
        certificado_id = str(uuid.uuid4())
        fecha_emision = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        self.cursor.execute('''
            INSERT INTO certificados (id, estudiante_id, curso_id, fecha_emision, archivo_certificado) 
            VALUES (?, ?, ?, ?, ?)
        ''', (certificado_id, estudiante[0], curso_id, fecha_emision, archivo_certificado))
        
        self.conn.commit()

    def __del__(self):
        if hasattr(self, 'conn'):
            self.conn.close()

def main():
    root = tk.Tk()
    app = UnexcaCertificateSystem(root)
    root.mainloop()

if __name__ == "__main__":
    main()